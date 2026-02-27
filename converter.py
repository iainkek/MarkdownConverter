"""
Markdown Converter — Bulk convert Word, Excel, PDF, EPUB, and URLs to Markdown.
"""

import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, scrolledtext
from pathlib import Path
from urllib.parse import urlparse

import windnd


# ─── Conversion Functions ───────────────────────────────────────────────────

def convert_docx(filepath):
    import mammoth
    with open(filepath, "rb") as f:
        result = mammoth.convert_to_html(f)
    from markdownify import markdownify
    return markdownify(result.value, heading_style="ATX")


def convert_xlsx(filepath):
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        parts.append(f"## {sheet_name}\n")
        # Build markdown table
        headers = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(headers) + " |")
        parts.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Pad or trim to match header count
            while len(cells) < len(headers):
                cells.append("")
            parts.append("| " + " | ".join(cells[:len(headers)]) + " |")
        parts.append("")
    wb.close()
    return "\n".join(parts)


def convert_pdf(filepath):
    import pdfplumber
    parts = []
    with pdfplumber.open(filepath) as pdf:
        for i, page in enumerate(pdf.pages, 1):
            text = page.extract_text()
            if text:
                parts.append(text)
            # Extract tables too
            tables = page.extract_tables()
            for table in tables:
                if not table or not table[0]:
                    continue
                headers = [str(c) if c else "" for c in table[0]]
                md_table = ["| " + " | ".join(headers) + " |"]
                md_table.append("| " + " | ".join(["---"] * len(headers)) + " |")
                for row in table[1:]:
                    cells = [str(c) if c else "" for c in row]
                    while len(cells) < len(headers):
                        cells.append("")
                    md_table.append("| " + " | ".join(cells[:len(headers)]) + " |")
                parts.append("\n".join(md_table))
    return "\n\n".join(parts)


def _find_calibre():
    """Find Calibre's ebook-convert executable."""
    import shutil

    # Check for bundled Calibre (installed by setup_app) via config file next to this exe
    if getattr(sys, 'frozen', False):
        app_dir = os.path.dirname(sys.executable)
    else:
        app_dir = os.path.dirname(os.path.abspath(__file__))
    config_file = os.path.join(app_dir, "calibre_path.txt")
    if os.path.isfile(config_file):
        try:
            calibre_dir = open(config_file, "r", encoding="utf-8").read().strip()
            exe = os.path.join(calibre_dir, "ebook-convert.exe")
            if os.path.isfile(exe):
                return exe
        except Exception:
            pass

    # Also check for Calibre Portable in same directory or parent
    for rel in ["Calibre Portable/Calibre", "Calibre", "../Calibre Portable/Calibre", "../Calibre"]:
        exe = os.path.join(app_dir, rel, "ebook-convert.exe")
        if os.path.isfile(exe):
            return exe

    # Check PATH
    path = shutil.which("ebook-convert")
    if path:
        return path

    # Common Windows install locations
    for base in [
        os.path.join(os.environ.get("PROGRAMFILES", "C:\\Program Files"), "Calibre2"),
        os.path.join(os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)"), "Calibre2"),
        os.path.join(os.path.expanduser("~"), "AppData", "Local", "calibre", "calibre"),
    ]:
        exe = os.path.join(base, "ebook-convert.exe")
        if os.path.isfile(exe):
            return exe
    return None


def _check_ade_activated():
    """Check if Adobe Digital Editions is activated on this machine (Windows registry)."""
    if sys.platform != "win32":
        return False, "ADE check only supported on Windows"
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Software\Adobe\Adept\Device")
        winreg.QueryValueEx(key, "key")
        winreg.CloseKey(key)
        return True, "ADE activated"
    except (FileNotFoundError, OSError):
        return False, (
            "Adobe Digital Editions is not activated on this computer.\n\n"
            "To fix this:\n"
            "1. Install Adobe Digital Editions from adobe.com/solutions/ebook/digital-editions\n"
            "2. Open it and sign in with your Adobe ID\n"
            "3. Open the EPUB file in ADE at least once\n"
            "4. Then try converting again here"
        )


def _epub_via_calibre(filepath):
    """Convert EPUB to text using Calibre. Uses calibredb add (triggers DeDRM) then ebook-convert."""
    import subprocess
    import tempfile
    import glob as globmod

    calibre_exe = _find_calibre()
    if not calibre_exe:
        return None

    calibre_dir = os.path.dirname(calibre_exe)
    calibredb = os.path.join(calibre_dir, "calibredb.exe")
    ebook_convert = calibre_exe  # ebook-convert.exe

    if not os.path.isfile(calibredb):
        return _epub_convert_direct(ebook_convert, filepath)

    # Point Calibre at the Portable config (where DeDRM plugin lives)
    env = os.environ.copy()
    settings_dir = os.path.normpath(os.path.join(calibre_dir, "..", "Calibre Settings"))
    if os.path.isdir(settings_dir):
        env["CALIBRE_CONFIG_DIRECTORY"] = settings_dir

    creation_flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0

    with tempfile.TemporaryDirectory() as tmpdir:
        lib_path = os.path.join(tmpdir, "lib")
        export_path = os.path.join(tmpdir, "export")
        os.makedirs(lib_path)
        os.makedirs(export_path)

        # Step 1: Import EPUB into temp Calibre library (DeDRM plugin runs here)
        try:
            result = subprocess.run(
                [calibredb, "add", str(filepath), "--library-path", lib_path, "-d"],
                capture_output=True, text=True, timeout=120,
                creationflags=creation_flags, env=env,
            )
        except subprocess.TimeoutExpired:
            return None
        except Exception:
            return None

        all_output = (result.stdout or "") + (result.stderr or "")

        # DeDRM logs to stdout — check for failure even if returncode is 0
        if "failed to decrypt" in all_output.lower() or "ADEPTError" in all_output:
            ade_ok, ade_msg = _check_ade_activated()
            if not ade_ok:
                raise ValueError(ade_msg)
            raise ValueError(
                "This EPUB is DRM-protected and DeDRM could not decrypt it.\n\n"
                "Adobe Digital Editions appears to be activated, but decryption "
                "still failed. Try opening the EPUB in ADE first, then convert again."
            )

        if result.returncode != 0:
            if "DRM" in all_output or "DeDRM" in all_output:
                ade_ok, ade_msg = _check_ade_activated()
                if not ade_ok:
                    raise ValueError(ade_msg)
                raise ValueError(
                    "This EPUB is DRM-protected and could not be decrypted. "
                    "Make sure Adobe Digital Editions is installed and authorized."
                )
            return None

        # Step 2: Export the (now decrypted) EPUB from the library
        try:
            result = subprocess.run(
                [calibredb, "export", "1", "--library-path", lib_path,
                 "--single-dir", "--to-dir", export_path, "--formats", "epub",
                 "--dont-save-cover", "--dont-write-opf", "--dont-save-extra-files"],
                capture_output=True, text=True, timeout=60,
                creationflags=creation_flags, env=env,
            )
        except Exception:
            return None

        # Find the exported EPUB
        exported = globmod.glob(os.path.join(export_path, "*.epub"))
        if not exported:
            return None

        # Step 3: Convert the decrypted EPUB → TXT via ebook-convert
        txt_path = os.path.join(tmpdir, "output.txt")
        try:
            result = subprocess.run(
                [ebook_convert, exported[0], txt_path, "--txt-output-encoding=utf-8"],
                capture_output=True, text=True, timeout=120,
                creationflags=creation_flags, env=env,
            )
            if result.returncode == 0 and os.path.isfile(txt_path):
                with open(txt_path, "r", encoding="utf-8", errors="replace") as f:
                    text = f.read().strip()
                if text:
                    return text
        except Exception:
            pass

    return None


def _epub_convert_direct(ebook_convert, filepath):
    """Fallback: convert EPUB directly with ebook-convert (no DRM support)."""
    import subprocess
    import tempfile

    creation_flags = subprocess.CREATE_NO_WINDOW if sys.platform == "win32" else 0
    with tempfile.TemporaryDirectory() as tmpdir:
        txt_path = os.path.join(tmpdir, "output.txt")
        try:
            result = subprocess.run(
                [ebook_convert, str(filepath), txt_path, "--txt-output-encoding=utf-8"],
                capture_output=True, text=True, timeout=120,
                creationflags=creation_flags,
            )
            if result.returncode == 0 and os.path.isfile(txt_path):
                with open(txt_path, "r", encoding="utf-8", errors="replace") as f:
                    text = f.read().strip()
                if text:
                    return text
        except Exception:
            pass
    return None


def _epub_via_zip(filepath):
    """Convert EPUB by extracting HTML directly from the zip (DRM-free only)."""
    import zipfile
    from markdownify import markdownify

    with zipfile.ZipFile(filepath, 'r') as zf:
        # Check for DRM encryption — if found, don't even try to read content
        if 'META-INF/encryption.xml' in zf.namelist():
            try:
                enc_xml = zf.read('META-INF/encryption.xml').decode('utf-8', errors='replace')
                if 'EncryptedData' in enc_xml:
                    return None  # DRM detected, signal to caller
            except Exception:
                pass

        # Find all HTML/XHTML files
        html_files = sorted([
            f for f in zf.namelist()
            if f.lower().endswith(('.html', '.xhtml', '.htm'))
            and 'toc' not in f.lower()
            and 'nav' not in f.lower()
        ])

        parts = []
        for html_file in html_files:
            try:
                raw = zf.read(html_file)
                # Must decode as valid UTF-8 (not latin-1 fallback, which masks encrypted data)
                html = raw.decode("utf-8")
                # Must contain actual HTML structure
                if '<!DOCTYPE' not in html[:500] and '<html' not in html[:500] and '<body' not in html[:1000]:
                    continue
                md = markdownify(html, heading_style="ATX")
                cleaned = md.strip()
                if cleaned:
                    parts.append(cleaned)
            except (UnicodeDecodeError, Exception):
                continue

    if parts:
        return "\n\n---\n\n".join(parts)
    return None


def convert_epub(filepath):
    # Try Calibre first (handles DRM if DeDRM plugin is installed)
    result = _epub_via_calibre(filepath)
    if result:
        return result

    # Fall back to direct zip extraction
    result = _epub_via_zip(filepath)
    if result:
        return result

    # Check if Calibre is available for the error message
    calibre = _find_calibre()
    if calibre:
        raise ValueError(
            "Could not extract content from this EPUB. "
            "It may be DRM-protected — make sure the DeDRM plugin is installed in Calibre."
        )
    raise ValueError(
        "This EPUB appears to be DRM-protected. "
        "Install Calibre (calibre-ebook.com) with the DeDRM plugin to convert it."
    )


def convert_url(url):
    import trafilatura
    downloaded = trafilatura.fetch_url(url)
    if not downloaded:
        raise ValueError(f"Could not fetch URL: {url}")
    result = trafilatura.extract(downloaded, output_format="markdown", include_tables=True,
                                  include_links=True, include_images=False)
    if not result:
        raise ValueError(f"Could not extract content from: {url}")
    return result


# ─── Helpers ────────────────────────────────────────────────────────────────

SUPPORTED_EXTENSIONS = {".docx", ".xlsx", ".pdf", ".epub"}

CONVERTERS = {
    ".docx": convert_docx,
    ".xlsx": convert_xlsx,
    ".pdf": convert_pdf,
    ".epub": convert_epub,
}


def safe_filename(name):
    """Sanitize a string into a safe filename."""
    name = re.sub(r'[<>:"/\\|?*]', '-', name)
    name = re.sub(r'\s+', '-', name)
    name = re.sub(r'-+', '-', name).strip('-')
    return name[:100] if name else "converted"


def unique_path(path):
    """Return a path that doesn't exist, appending _1, _2, etc. if needed."""
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 1
    while True:
        new_path = parent / f"{stem}_{counter}{suffix}"
        if not new_path.exists():
            return new_path
        counter += 1


def url_to_filename(url):
    """Generate a markdown filename from a URL."""
    parsed = urlparse(url)
    domain = parsed.netloc.replace("www.", "")
    path_part = parsed.path.strip("/").replace("/", "-")
    name = f"{domain}-{path_part}" if path_part else domain
    return safe_filename(name)


# ─── GUI ────────────────────────────────────────────────────────────────────

class MarkdownConverterApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Markdown Converter")
        self.root.geometry("700x650")
        self.root.minsize(600, 500)
        self.root.configure(bg="#f0f0f0")

        self.files = []
        self.converting = False

        self._build_ui()
        self._setup_drag_and_drop()

    def _build_ui(self):
        root = self.root

        # Title
        title = tk.Label(root, text="Markdown Converter", font=("Segoe UI", 18, "bold"),
                         bg="#f0f0f0", fg="#333")
        title.pack(pady=(15, 5))

        subtitle = tk.Label(root, text="Convert Word, Excel, PDF, EPUB & URLs to Markdown",
                            font=("Segoe UI", 10), bg="#f0f0f0", fg="#666")
        subtitle.pack(pady=(0, 10))

        # Drop zone
        drop_frame = tk.Frame(root, bg="#f0f0f0")
        drop_frame.pack(fill=tk.X, padx=20)

        self.drop_zone = tk.Label(
            drop_frame,
            text="Drag & drop files here\nor click to browse\n\nSupports: .docx  .xlsx  .pdf  .epub",
            font=("Segoe UI", 11),
            bg="white", fg="#888",
            relief="solid", borderwidth=1,
            height=6, cursor="hand2"
        )
        self.drop_zone.pack(fill=tk.X)
        self.drop_zone.bind("<Button-1>", self._browse_files)

        # File list
        self.file_list_var = tk.StringVar(value="No files selected")
        self.file_list_label = tk.Label(drop_frame, textvariable=self.file_list_var,
                                         font=("Segoe UI", 9), bg="#f0f0f0", fg="#555",
                                         anchor="w", justify="left", wraplength=650)
        self.file_list_label.pack(fill=tk.X, pady=(5, 0))

        # Clear files button
        self.clear_btn = tk.Button(drop_frame, text="Clear files", font=("Segoe UI", 9),
                                    command=self._clear_files, state=tk.DISABLED)
        self.clear_btn.pack(anchor="e", pady=(2, 0))

        # URL section
        url_frame = tk.Frame(root, bg="#f0f0f0")
        url_frame.pack(fill=tk.X, padx=20, pady=(10, 0))

        url_label = tk.Label(url_frame, text="Website URLs (one per line):",
                              font=("Segoe UI", 10), bg="#f0f0f0", fg="#333")
        url_label.pack(anchor="w")

        self.url_text = scrolledtext.ScrolledText(url_frame, height=4, font=("Segoe UI", 10),
                                                    relief="solid", borderwidth=1)
        self.url_text.pack(fill=tk.X)

        # Convert button
        btn_frame = tk.Frame(root, bg="#f0f0f0")
        btn_frame.pack(fill=tk.X, padx=20, pady=(15, 0))

        self.convert_btn = tk.Button(
            btn_frame, text="Convert to Markdown",
            font=("Segoe UI", 12, "bold"),
            bg="#4CAF50", fg="white", activebackground="#45a049",
            cursor="hand2", relief="flat", padx=20, pady=8,
            command=self._start_conversion
        )
        self.convert_btn.pack()

        # Progress
        self.progress_var = tk.StringVar(value="Ready")
        self.progress_label = tk.Label(btn_frame, textvariable=self.progress_var,
                                        font=("Segoe UI", 10), bg="#f0f0f0", fg="#555")
        self.progress_label.pack(pady=(5, 0))

        self.progress_bar = ttk.Progressbar(btn_frame, mode="determinate")
        self.progress_bar.pack(fill=tk.X, pady=(5, 0))

        # Results
        results_frame = tk.Frame(root, bg="#f0f0f0")
        results_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(10, 15))

        results_label = tk.Label(results_frame, text="Results:",
                                  font=("Segoe UI", 10), bg="#f0f0f0", fg="#333")
        results_label.pack(anchor="w")

        self.results_text = scrolledtext.ScrolledText(results_frame, height=8,
                                                        font=("Consolas", 9),
                                                        relief="solid", borderwidth=1,
                                                        state=tk.DISABLED)
        self.results_text.pack(fill=tk.BOTH, expand=True)
        self.results_text.tag_configure("success", foreground="#2e7d32")
        self.results_text.tag_configure("error", foreground="#c62828")
        self.results_text.tag_configure("info", foreground="#1565c0")

    def _setup_drag_and_drop(self):
        windnd.hook_dropfiles(self.root, func=self._on_drop)

    def _on_drop(self, file_list):
        new_files = []
        for f in file_list:
            path = f.decode("utf-8") if isinstance(f, bytes) else str(f)
            p = Path(path)
            if p.suffix.lower() in SUPPORTED_EXTENSIONS:
                new_files.append(p)
        if new_files:
            self.files.extend(new_files)
            self._update_file_list()

    def _browse_files(self, event=None):
        filetypes = [
            ("Supported files", "*.docx *.xlsx *.pdf *.epub"),
            ("Word documents", "*.docx"),
            ("Excel spreadsheets", "*.xlsx"),
            ("PDF files", "*.pdf"),
            ("EPUB books", "*.epub"),
            ("All files", "*.*"),
        ]
        paths = filedialog.askopenfilenames(filetypes=filetypes)
        if paths:
            for p in paths:
                pp = Path(p)
                if pp.suffix.lower() in SUPPORTED_EXTENSIONS:
                    self.files.append(pp)
            self._update_file_list()

    def _update_file_list(self):
        if self.files:
            names = [f"  {f.name}" for f in self.files]
            self.file_list_var.set(f"{len(self.files)} file(s) selected:\n" + "\n".join(names))
            self.clear_btn.configure(state=tk.NORMAL)
        else:
            self.file_list_var.set("No files selected")
            self.clear_btn.configure(state=tk.DISABLED)

    def _clear_files(self):
        self.files.clear()
        self._update_file_list()

    def _log(self, msg, tag="info"):
        self.results_text.configure(state=tk.NORMAL)
        self.results_text.insert(tk.END, msg + "\n", tag)
        self.results_text.see(tk.END)
        self.results_text.configure(state=tk.DISABLED)

    def _start_conversion(self):
        if self.converting:
            return

        urls_text = self.url_text.get("1.0", tk.END).strip()
        urls = [u.strip() for u in urls_text.splitlines() if u.strip()]

        if not self.files and not urls:
            self.progress_var.set("Nothing to convert — add files or URLs first")
            return

        self.converting = True
        self.convert_btn.configure(state=tk.DISABLED, bg="#999")
        self.results_text.configure(state=tk.NORMAL)
        self.results_text.delete("1.0", tk.END)
        self.results_text.configure(state=tk.DISABLED)

        thread = threading.Thread(target=self._do_conversion, args=(list(self.files), urls), daemon=True)
        thread.start()

    def _do_conversion(self, files, urls):
        total = len(files) + len(urls)
        done = 0

        self.root.after(0, lambda: self.progress_bar.configure(maximum=total, value=0))
        self.root.after(0, lambda: self.progress_var.set(f"Converting 0/{total}..."))

        # Convert files
        for filepath in files:
            ext = filepath.suffix.lower()
            converter = CONVERTERS.get(ext)
            if not converter:
                self.root.after(0, lambda fp=filepath: self._log(f"  SKIP  {fp.name} — unsupported type", "error"))
                done += 1
                continue

            self.root.after(0, lambda fp=filepath: self._log(f"  ...   {fp.name}", "info"))
            try:
                markdown = converter(str(filepath))
                out_path = unique_path(filepath.with_suffix(".md"))
                out_path.write_text(markdown, encoding="utf-8")
                self.root.after(0, lambda fp=filepath, op=out_path: self._log(
                    f"  OK    {fp.name}  →  {op.name}", "success"))
            except Exception as e:
                self.root.after(0, lambda fp=filepath, err=e: self._log(
                    f"  FAIL  {fp.name}  —  {err}", "error"))

            done += 1
            d = done
            self.root.after(0, lambda d=d: self.progress_bar.configure(value=d))
            self.root.after(0, lambda d=d, t=total: self.progress_var.set(f"Converting {d}/{t}..."))

        # Convert URLs
        for url in urls:
            self.root.after(0, lambda u=url: self._log(f"  ...   {u}", "info"))
            try:
                markdown = convert_url(url)
                fname = url_to_filename(url) + ".md"
                out_dir = Path.home() / "Documents"
                if not out_dir.exists():
                    out_dir = Path.cwd()
                out_path = unique_path(out_dir / fname)
                out_path.write_text(markdown, encoding="utf-8")
                self.root.after(0, lambda u=url, op=out_path: self._log(
                    f"  OK    {u}  →  {op}", "success"))
            except Exception as e:
                self.root.after(0, lambda u=url, err=e: self._log(
                    f"  FAIL  {u}  —  {err}", "error"))

            done += 1
            d = done
            self.root.after(0, lambda d=d: self.progress_bar.configure(value=d))
            self.root.after(0, lambda d=d, t=total: self.progress_var.set(f"Converting {d}/{t}..."))

        self.root.after(0, lambda: self.progress_var.set(f"Done — {total} item(s) processed"))
        self.root.after(0, lambda: self.convert_btn.configure(state=tk.NORMAL, bg="#4CAF50"))
        self.root.after(0, self._clear_after_conversion)
        self.converting = False

    def _clear_after_conversion(self):
        self.files.clear()
        self._update_file_list()

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = MarkdownConverterApp()
    app.run()
